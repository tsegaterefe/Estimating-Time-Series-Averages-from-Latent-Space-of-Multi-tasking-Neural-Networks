
"""
Created on Fri Feb 28 09:27:22 2020

@author: Tsega
"""
import Warping_Reg_Net as encoder
import Warping_Reg_Net_Config as my_configs
import time
import One_NN as myNN
import numpy as npy
import pandas as pd
from importlib import reload  
import openpyxl as write_results_to_excel
import os
def main():
    my_config=my_configs.Warping_Reg_Net_Config()
    Files_to_execute_for=pd.read_csv(my_config.List_of_data_sets+my_config.List_of_data_sets_FName, sep=',',header=None)
    start_execution_from=pd.read_csv(my_config.To_start_from+my_config.To_start_from_FName, sep=',',header=None)
    start_execution_at=pd.read_csv(my_config.To_start_from+my_config.start_at, sep=',',header=None)
    root=my_config.File_loc
    model_save_root=my_config.Model_save_path
    start_time=0
    end_time=0
    for i in range(Files_to_execute_for.shape[0]):
        temp=Files_to_execute_for.iloc[i]
        temp=temp.to_string(index=False)
        temp=temp.replace(' ','')
        my_config.File_loc=root+temp+'/'
        my_config.File_name=temp
        my_config.Model_save_name=temp+'.ckpt'
        my_config.Model_save_path=model_save_root
        print('path exists',os.path.isdir(my_config.Model_save_path+my_config.File_name))
        if my_config.First_time_train==0 and os.path.isdir(my_config.Model_save_path+my_config.File_name)==False:
            break
        try:
            os.mkdir(my_config.Model_save_path+my_config.File_name)
        except OSError:
            print ("Creation of the directory %s failed" % my_config.Model_save_path+my_config.File_name)
        else:
            print ("Successfully created the directory %s " % my_config.Model_save_path+my_config.File_name)
        for k in range(int(start_execution_from.iloc[i]),my_config.run_per_single_data):
            for j in range(int(start_execution_at.iloc[i]),len(my_config.Quantiles)):
                my_config.Model_save_path=model_save_root
                try:
                    os.mkdir(my_config.Model_save_path+my_config.File_name+'/'+'Trial'+str(k)+str(j+1))
                except OSError:
                    print ("Creation of the directory %s failed" % my_config.Model_save_path+my_config.File_name+'/'+'Trial'+str(j+1))
                else:
                    print ("Successfully created the directory %s " %my_config.Model_save_path+my_config.File_name+'/'+'Trial'+str(j+1))
                my_config.Model_save_path=my_config.Model_save_path+my_config.File_name+'/'+'Trial'+str(k)+str(j+1)+'/'
                print('Analyzing data sets........................')
                mydatas=encoder.Warping_Reg_Net(my_config.File_loc,my_config.File_name,my_config.Model_save_path, my_config.Model_save_name,my_config.classif_mode,my_config.select_all,my_config.class_label,my_config.validation_size)
                print('Total data sets in train file:',mydatas.temp.shape[0],mydatas.temp.shape[1])
                print('Total data sets in test file:',mydatas.temps.shape[0],mydatas.temp.shape[1])
                print('Unique classes:', mydatas.uniques)
                print('Total members of classes in the concatinated data set:', mydatas.counts)
                print('Finished analyzing data sets.')
                my_config.encoder_node_size=int((mydatas.temp.shape[1]-1)/4)
                if my_config.First_time_train==1:
                    if my_config.model_type==1:
                        mydatas.Multi_task_encoder_two(my_config.Quantiles[j],my_config.encoder_node_size,my_config.filter_size,my_config.polling_size,my_config.En_L1_reg,my_config.En_L2_reg,my_config.De_L1_reg,my_config.De_L2_reg,my_config.Cl_L1_reg,my_config.Cl_L2_reg,my_config.Input_activ,my_config.Hidden_activ,my_config.Learning_rate)
                        start_time=time.time()
                        mydatas.Train_Warper_Model(my_config.Epoch,my_config.batch_size)
                    else:
                        if my_config.model_type==2:
                            mydatas.Build_Inception_Multitask(my_config.Quantiles[j],my_config.encoder_node_size,my_config.filter_size,my_config.polling_size,my_config.En_L1_reg,my_config.En_L2_reg,my_config.De_L1_reg,my_config.De_L2_reg,my_config.Cl_L1_reg,my_config.Cl_L2_reg,my_config.Input_activ,my_config.Hidden_activ,my_config.Learning_rate)
                            start_time=time.time()
                            mydatas.Train_Warper_Model(my_config.Epoch,my_config.batch_size)
                        else:
                            if my_config.model_type==3:
                                mydatas.Build_VGG_ResNet_Multitasking(my_config.Quantiles[j],my_config.encoder_node_size,my_config.filter_size,my_config.polling_size,my_config.En_L1_reg,my_config.En_L2_reg,my_config.De_L1_reg,my_config.De_L2_reg,my_config.Cl_L1_reg,my_config.Cl_L2_reg,my_config.Input_activ,my_config.Hidden_activ,my_config.Learning_rate)
                                start_time=time.time()
                                mydatas.Train_Warper_Model(my_config.Epoch,my_config.batch_size)
                else:
                    if my_config.model_type==1:
                            mydatas.Multi_task_encoder_two(my_config.Quantiles[j],my_config.encoder_node_size,my_config.filter_size,my_config.polling_size,my_config.En_L1_reg,my_config.En_L2_reg,my_config.De_L1_reg,my_config.De_L2_reg,my_config.Cl_L1_reg,my_config.Cl_L2_reg,my_config.Input_activ,my_config.Hidden_activ,my_config.Learning_rate)
                            mydatas.load_model_weights(my_config.Model_save_name,my_config.Model_save_path)
                    else:
                        if my_config.model_type==2:
                                mydatas.Build_Inception_Multitask(my_config.Quantiles[j],my_config.encoder_node_size,my_config.filter_size,my_config.polling_size,my_config.En_L1_reg,my_config.En_L2_reg,my_config.De_L1_reg,my_config.De_L2_reg,my_config.Cl_L1_reg,my_config.Cl_L2_reg,my_config.Input_activ,my_config.Hidden_activ,my_config.Learning_rate)
                                mydatas.load_model_weights(my_config.Model_save_name,my_config.Model_save_path)
                        else:
                                mydatas.Build_VGG_ResNet_Multitasking(my_config.Quantiles[j],my_config.encoder_node_size,my_config.filter_size,my_config.polling_size,my_config.En_L1_reg,my_config.En_L2_reg,my_config.De_L1_reg,my_config.De_L2_reg,my_config.Cl_L1_reg,my_config.Cl_L2_reg,my_config.Input_activ,my_config.Hidden_activ,my_config.Learning_rate)
                                mydatas.load_model_weights(my_config.Model_save_name,my_config.Model_save_path)
                if my_config.classif_mode==0:
                    end_time=time.time()
                    my_latents,my_lat_means,time_estimate,lat_wgss,time_wgss=mydatas.fit_Gausian_Latent_two()
                    mydatas.write_means(0,end_time-start_time,' time taken to estimate mean.tsv')
                    mydatas.write_means()
                    mydatas.write_atsv_file(my_latents,my_config.Model_save_path,'Train latent representation.tsv')
                    display={'Data_Set':[''],'Time_Domain_DTW_Classification':[0],'Time_Domain_Euclidean_Classification':[0],'Latent_Space_Classification':[0],'Training_Time':[0],'Trial_No':[0],'latent_Euclidean_WGSS':[''],'latent_DTW_WGSS':[''],'Time_Euclidean_WGSS':[''],'Time_DTW_WGSS':['']}
                    if my_config.select_all==0 and my_config.classif_mode==0 and npy.isnan(my_lat_means).any()==False and npy.isnan(time_estimate).any()==False:
                        est_data,my_test_mean,my_test,decoded=mydatas.estimate_mean_separately()
                        mydatas.write_atsv_file(est_data,my_config.Model_save_path,'Test latent space.tsv')
                        print('Beginning Time Domain DTW Classification........')
                        myNN_classif=myNN.One_NN(time_estimate,'null','null',0,data=my_test)
                        res=myNN_classif.classify_DTW()
                        display['Time_Domain_DTW_Classification']=[res]
                        print('Beginning Time Domain Euclidean Classification........')
                        myNN_classif=myNN.One_NN(time_estimate,'null','null',0,data=my_test)
                        res=myNN_classif.classify_Euclidean()
                        display['Time_Domain_Euclidean_Classification']=[res]
                        print('Begnining Latent Space Classification........')
                        myNN_classif=myNN.One_NN(my_lat_means,'null','null',0,data=est_data)
                        res=myNN_classif.classify_Euclidean()
                        display['Latent_Space_Classification']=[res]
                        print("Time Domain DTW Results:",display['Time_Domain_DTW_Classification'],"Time Domain Euclidean Results:",display['Time_Domain_Euclidean_Classification'],"Latent doamin result:",display['Latent_Space_Classification'])
                        tsvdata=npy.array([display['Time_Domain_DTW_Classification'],display['Latent_Space_Classification']])
                        mydatas.write_atsv_file(tsvdata.reshape((1,2)),my_config.Model_save_path,'Time domain Latent classification result.tsv')
                        display['Training_Time']=[end_time-start_time]
                        display['Trial_No']=[j]
                        display['latent_Euclidean_WGSS']=[str(lat_wgss[0,:])]
                        display['latent_DTW_WGSS']=[str(lat_wgss[1,:])]
                        display['Time_Euclidean_WGSS']=[str(time_wgss[0,:])]
                        display['Time_DTW_WGSS']=[str(time_wgss[1,:])]
                        df=pd.DataFrame(display)
                        if os.path.isfile(model_save_root+'Encoder_Decoder_statistical_acess_quantile.xlsx'):
                            display['Data_Set']=my_config.File_name
                            print('my config file name',my_config.File_name)
                            print(display['Data_Set'])
                            wb=write_results_to_excel.load_workbook(model_save_root+'Encoder_Decoder_statistical_acess_quantile.xlsx',read_only=False)
                            sheets=wb.sheetnames
                            main_sheet=wb[sheets[0]]
                            lastrow=main_sheet.max_row
                            indexvalue=main_sheet.cell(row=lastrow,column=10).value
                            while indexvalue is None and lastrow > 0:
                                    lastrow -= 1
                            indexvalue=main_sheet.cell(row=lastrow,column=1).value
                            main_sheet.cell(row=lastrow+1,column=1,value=display['Data_Set'])
                            main_sheet.cell(row=lastrow+1,column=2,value=display['Time_Domain_DTW_Classification'][0])
                            main_sheet.cell(row=lastrow+1,column=3,value=display['Time_Domain_Euclidean_Classification'][0])
                            main_sheet.cell(row=lastrow+1,column=4,value=display['Latent_Space_Classification'][0])
                            main_sheet.cell(row=lastrow+1,column=5,value=display['Training_Time'][0])
                            main_sheet.cell(row=lastrow+1,column=6,value=display['Trial_No'][0])
                            main_sheet.cell(row=lastrow+1,column=7,value=display['latent_Euclidean_WGSS'][0])
                            main_sheet.cell(row=lastrow+1,column=8,value=display['latent_DTW_WGSS'][0])
                            main_sheet.cell(row=lastrow+1,column=9,value=display['Time_Euclidean_WGSS'][0])
                            main_sheet.cell(row=lastrow+1,column=10,value=display['Time_DTW_WGSS'][0])
                            wb.save(model_save_root+'Encoder_Decoder_statistical_acess_quantile.xlsx')
                            wb.close
                        else:
                            print('my config file name',my_config.File_name)
                            display['Data_Set']=[str(my_config.File_name)]
                            df=pd.DataFrame(display)
                            df.to_excel(model_save_root+'Encoder_Decoder_statistical_acess_quantile.xlsx',sheet_name='Results',index=False)
                reload(encoder)
                reload(myNN)
                start_execution_at.iloc[i,0]=start_execution_at.iloc[i,0]+1
                start_execution_at.to_csv(my_config.To_start_from+my_config.start_at,index=False,header=None)
            start_execution_at.iloc[i,0]=0
            start_execution_at.to_csv(my_config.To_start_from+my_config.start_at,index=False,header=None)    
            start_execution_from.iloc[i,0]=start_execution_from.iloc[i,0]+1
            start_execution_from.to_csv(my_config.To_start_from+my_config.To_start_from_FName,index=False,header=None)
        start_execution_from.iloc[i+1:,0].to_csv(my_config.To_start_from+my_config.To_start_from_FName,index=False,header=None)
        Files_to_execute_for.iloc[i+1:,0].to_csv(my_config.List_of_data_sets+my_config.List_of_data_sets_FName,index=False,header=None)
        start_execution_at.iloc[i+1:,0].to_csv(my_config.To_start_from+my_config.start_at,index=False,header=None) 
if __name__ == "__main__":
        main()
